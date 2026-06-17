import os
import sys

# Setup Django
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'dairypro.settings')
sys.path.insert(0, os.path.dirname(__file__))

import django
django.setup()

import openpyxl
from decimal import Decimal, InvalidOperation
from datetime import date

from dairypro.core.models import User
from dairypro.cattle.models import Cattle, CattleStatus
from dairypro.milk.models import MilkCollection, Shift

print("=" * 60)
print("  DairyPro 360 - Milk Production Data Import")
print("=" * 60)

EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'milk_import.xlsx')
admin = User.objects.filter(is_superuser=True).first()

VALID_COWS = {
    'Shobha', 'Devaki', 'Sita', 'Lakshmi', 'Mruga', 'Radha',
    'Shravani', 'Uma', 'Parvati', 'Durga', 'Ganga', 'Yamuna',
    'Kapila', 'Sindhu', 'Pushpa'
}

DATA_SHEETS = [
    'June', 'July', 'August', 'September ', 'October ',
    ' November ', 'December', 'January 2026', 'Feb 2026',
    'Mar 2026', 'April 2026', 'May 2026', 'June 2026'
]

# Step 1: Create cattle
print("\n[1/3] Creating cattle records...")
cattle_map = {}
for cow_name in sorted(VALID_COWS):
    tag = 'COW-' + cow_name.upper().replace(' ', '')[:8]
    obj, created = Cattle.objects.get_or_create(
        name=cow_name,
        defaults={
            'tag_number': tag,
            'breed': 'Indigenous',
            'status': CattleStatus.LACTATING,
            'is_active': True,
            'created_by': admin,
            'date_of_birth': date(2018, 1, 1),
        }
    )
    cattle_map[cow_name] = obj
    status = 'Created' if created else 'Found  '
    print("  " + status + ": " + cow_name + " (" + obj.tag_number + ")")

print("  Total cattle: " + str(len(cattle_map)))

# Step 2: Parse Excel
print("\n[2/3] Parsing Excel...")
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
records = []

for sheet_name in DATA_SHEETS:
    if sheet_name not in wb.sheetnames:
        print("  SKIP (not found): " + sheet_name)
        continue

    ws = wb[sheet_name]
    header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]

    cow_cols = {}
    for i, h in enumerate(header):
        if h and isinstance(h, str) and h.strip() in VALID_COWS:
            cow_cols[i] = h.strip()

    if not cow_cols:
        print("  SKIP (no cow cols): " + sheet_name)
        continue

    current_date = None
    sheet_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        cell_a = row[0]
        cell_b = row[1]

        if not cell_b or str(cell_b).strip() not in ('Morning', 'Evening'):
            continue

        if cell_a is not None:
            if hasattr(cell_a, 'date'):
                current_date = cell_a.date()
            elif isinstance(cell_a, date):
                current_date = cell_a

        if current_date is None:
            continue

        shift = Shift.MORNING if str(cell_b).strip() == 'Morning' else Shift.EVENING

        for col_idx, cow_name in cow_cols.items():
            if col_idx >= len(row):
                continue
            val = row[col_idx]
            if val is None or val == '':
                continue
            if isinstance(val, str) and (val.startswith('=') or val.strip() == ''):
                continue
            try:
                qty = Decimal(str(val)).quantize(Decimal('0.01'))
                if qty <= Decimal('0'):
                    continue
            except (InvalidOperation, ValueError):
                continue

            records.append({
                'cattle': cattle_map[cow_name],
                'date': current_date,
                'shift': shift,
                'qty': qty,
            })
            sheet_count += 1

    print("  Parsed '" + sheet_name.strip() + "': " + str(sheet_count) + " entries")

print("  Total records to insert: " + str(len(records)))

# Step 3: Insert
print("\n[3/3] Inserting into database...")
inserted = 0
duplicate = 0
errors = 0

for r in records:
    try:
        obj, created = MilkCollection.objects.get_or_create(
            cattle=r['cattle'],
            collection_date=r['date'],
            shift=r['shift'],
            defaults={
                'field_worker': admin,
                'quantity_litres': r['qty'],
                'fat_percentage': Decimal('4.0'),
                'snf_percentage': Decimal('8.5'),
            }
        )
        if created:
            inserted += 1
        else:
            duplicate += 1
    except Exception as e:
        errors += 1
        print("  ERROR: " + str(e))

print("  Inserted : " + str(inserted))
print("  Duplicate: " + str(duplicate))
print("  Errors   : " + str(errors))
print()
print("=" * 60)
print("  Done! " + str(inserted) + " milk records imported.")
print("  Visit http://localhost:3000 -> Milk Collection")
print("=" * 60)
